from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.chart import (LineChart, BarChart, PieChart, Reference, Series, reference)
from openpyxl.chart.label import DataLabelList
import tkinter.messagebox
import Data

wb = Workbook()

def ExcelGraph():
    A()
    B()
    C()
    if Data.rb > 1:
        D()
    E()
    
def A():
    sheetA = wb.create_sheet('전체 출원동향', 0)
    A그래프data = Data.전체출원동향()

    for r in dataframe_to_rows(A그래프data, index=False, header=True):
        sheetA.append(r)

    sheetA.insert_cols(2)
    for row, cellobj in enumerate(list(sheetA.columns)[1]):
        n = '=right(A%d,2)' % (row+1)
        cellobj.value = n

    chartA1 = BarChart()
    dataA1 = Reference(sheetA, min_col = 3, min_row = 1, max_row = 21)
    catsA1 = Reference(sheetA, min_col = 2, min_row = 2, max_row = 21)
    chartA1.add_data(dataA1, titles_from_data = True)
    chartA1.set_categories(catsA1)
    chartA1.y_axis.majorGridlines = None

    chartA2 = LineChart()
    dataA2 = Reference(sheetA, min_col = 4, min_row = 1, max_row = 21)
    chartA2.add_data(dataA2, titles_from_data = True)
    chartA2.y_axis.majorGridlines = None
    chartA2.y_axis.axId = 2000

    # y축 위치 변경
    chartA2.y_axis.crosses = 'max'
    # 그래프 합치기
    chartA1 += chartA2
    chartA1.width = 15
    chartA1.height = 10
    chartA1.legend.position = 't'
    chartA1.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True)) # 테두리 제거
    sheetA.add_chart(chartA1, 'F2')

    global savepath
    savepath = Data.Save()

def B():
    sheetB = wb.create_sheet('주요국가별 출원동향', 1)
    B그래프data = Data.주요국출원동향()

    for r in dataframe_to_rows(B그래프data, index=False, header=True):
        sheetB.append(r)

    sheetB.insert_cols(2)
    for row, cellobj in enumerate(list(sheetB.columns)[1]):
        n = '=right(A%d,2)' % (row+1)
        cellobj.value = n

    sheetB['B22'] = '합계'
    sheetB['C22'] = '=SUM(C2:C21)'
    sheetB['D22'] = '=SUM(D2:D21)'
    sheetB['E22'] = '=SUM(E2:E21)'
    sheetB['F22'] = '=SUM(F2:F21)'

    chartB1 = LineChart()
    dataB1 = Reference(sheetB, min_col = 3, min_row = 1, max_row = 21, max_col = 6)
    catsB1 = Reference(sheetB, min_col = 2, min_row = 2, max_row = 21)
    chartB1.add_data(dataB1, titles_from_data = True)
    chartB1.set_categories(catsB1)
    chartB1.y_axis.majorGridlines = None
    chartB1.width = 15
    chartB1.height = 10
    chartB1.legend.position = 't'
    chartB1.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    sheetB.add_chart(chartB1, 'H2')

    chartB2 = PieChart()
    dataB2 = Reference(sheetB, min_col = 3, min_row = 22, max_col = 6)
    labelsB2 = Reference(sheetB, min_col = 3, min_row = 1, max_col = 6)
    chartB2.add_data(dataB2, from_rows = 22, titles_from_data = False)
    chartB2.set_categories(labelsB2)
    chartB2.width = 5
    chartB2.height =5
    chartB2.legend = None
    chartB2.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chartB2.dLbls = DataLabelList()
    chartB2.dLbls.showPercent = True

    sheetB.add_chart(chartB2, 'M1')

def C():
    # 데이터 정리
    sheetC = wb.create_sheet('주요국 내 상위다출원국가', 2)
    C상위다출원국가KR = Data.상위다출원국가('KR')
    C상위다출원국가JP = Data.상위다출원국가('JP')
    C상위다출원국가US = Data.상위다출원국가('US')
    C상위다출원국가EP = Data.상위다출원국가('EP')

    for r in dataframe_to_rows(C상위다출원국가KR, index=False, header=True):
        sheetC.append(r)

    sheetC['A22'] = ''
    sheetC['A23'] = ''

    for r in dataframe_to_rows(C상위다출원국가JP, index=False, header=True):
        sheetC.append(r)

    sheetC['A45'] = ''
    sheetC['A46'] = ''

    for r in dataframe_to_rows(C상위다출원국가US, index=False, header=True):
        sheetC.append(r)

    sheetC['A68'] = ''
    sheetC['A69'] = ''

    for r in dataframe_to_rows(C상위다출원국가EP, index=False, header=True):
        sheetC.append(r)

    sheetC['A91'] = ''
    sheetC['A92'] = ''

    sheetC.insert_cols(2)
    for row, cellobj in enumerate(list(sheetC.columns)[1]):
        n = '=right(A%d,2)' % (row+1)
        cellobj.value = n

    # 엑셀 함수
    sheetC['B22'] = '합계'
    sheetC['C22'] = '=SUM(C2:C21)'
    sheetC['D22'] = '=SUM(D2:D21)'
    sheetC['E22'] = '=SUM(E2:E21)'
    sheetC['F22'] = '=SUM(F2:F21)'
    sheetC['A23'] = ' '

    sheetC['B44'] = '합계'
    sheetC['C45'] = '=SUM(C25:C44)'
    sheetC['D45'] = '=SUM(D25:D44)'
    sheetC['E45'] = '=SUM(E25:E44)'
    sheetC['F45'] = '=SUM(F25:F44)'
    sheetC['A46'] = ' '

    sheetC['B68'] = '합계'
    sheetC['C68'] = '=SUM(C48:C67)'
    sheetC['D68'] = '=SUM(D48:D67)'
    sheetC['E68'] = '=SUM(E48:E67)'
    sheetC['F68'] = '=SUM(F48:F67)'
    sheetC['A69'] = ' '

    sheetC['B91'] = '합계'
    sheetC['C91'] = '=SUM(C71:C90)'
    sheetC['D91'] = '=SUM(D71:D90)'
    sheetC['E91'] = '=SUM(E71:E90)'
    sheetC['F91'] = '=SUM(F71:F90)'
    sheetC['A92'] = ' '

    # 그래프 그리기
    chartC11 = LineChart()
    dataC11 = Reference(sheetC, min_col = 3, min_row = 1, max_col = 6, max_row = 21)
    catsC11 = Reference(sheetC, min_col = 2, min_row = 2, max_row = 21)
    chartC11.add_data(dataC11, titles_from_data = True)
    chartC11.set_categories(catsC11)
    chartC11.y_axis.majorGridlines = None
    chartC11.width = 15
    chartC11.height = 10
    chartC11.legend.position = 't'
    chartC11.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    sheetC.add_chart(chartC11, 'H2')

    chartC12 = PieChart()
    dataC12 = Reference(sheetC, min_col = 3, min_row = 22, max_col = 6)
    labelC12 = Reference(sheetC, min_col = 3, min_row = 1, max_col = 6)
    chartC12.add_data(dataC12, from_rows = 22, titles_from_data = False)
    chartC12.set_categories(labelC12)
    chartC12.width = 5
    chartC12.height = 5
    chartC12.legend = None
    chartC12.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chartC12.dLbls = DataLabelList()
    chartC12.dLbls.showPercent = True
    sheetC.add_chart(chartC12, 'M1')

    chartC21 = LineChart()
    dataC21 = Reference(sheetC, min_col = 3, min_row = 24, max_col = 6, max_row = 44)
    catsC21 = Reference(sheetC, min_col = 2, min_row = 25, max_row = 44)
    chartC21.add_data(dataC21, titles_from_data = True)
    chartC21.set_categories(catsC21)
    chartC21.y_axis.majorGridlines = None
    chartC21.width = 15
    chartC21.height = 10
    chartC21.legend.position = 't'
    chartC21.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    sheetC.add_chart(chartC21, 'H25')

    chartC22 = PieChart()
    dataC22 = Reference(sheetC, min_col = 3, min_row = 45, max_col = 6)
    labelC22 = Reference(sheetC, min_col = 3, min_row = 24, max_col = 6)
    chartC22.add_data(dataC22, from_rows = 45, titles_from_data = False)
    chartC22.set_categories(labelC22)
    chartC22.width = 5
    chartC22.height = 5
    chartC22.legend = None
    chartC22.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chartC22.dLbls = DataLabelList()
    chartC22.dLbls.showPercent = True
    sheetC.add_chart(chartC22, 'M24')

    chartC31 = LineChart()
    dataC31 = Reference(sheetC, min_col = 3, min_row = 47, max_col = 6, max_row = 67)
    catsC31 = Reference(sheetC, min_col = 2, min_row = 48, max_row = 67)
    chartC31.add_data(dataC31, titles_from_data = True)
    chartC31.set_categories(catsC31)
    chartC31.y_axis.majorGridlines = None
    chartC31.width = 15
    chartC31.height = 10
    chartC31.legend.position = 't'
    chartC31.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    sheetC.add_chart(chartC31, 'H48')

    chartC32 = PieChart()
    dataC32 = Reference(sheetC, min_col = 3, min_row = 68, max_col = 6)
    labelC32 = Reference(sheetC, min_col = 3, min_row = 47, max_col = 6)
    chartC32.add_data(dataC32, from_rows = 68, titles_from_data = False)
    chartC32.set_categories(labelC32)
    chartC32.width = 5
    chartC32.height = 5
    chartC32.legend = None
    chartC32.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chartC32.dLbls = DataLabelList()
    chartC32.dLbls.showPercent = True
    sheetC.add_chart(chartC32, 'M47')

    chartC41 = LineChart()
    dataC41 = Reference(sheetC, min_col = 3, min_row = 70, max_col = 6, max_row = 90)
    catsC41 = Reference(sheetC, min_col = 2, min_row = 71, max_row = 90)
    chartC41.add_data(dataC41, titles_from_data = True)
    chartC41.set_categories(catsC41)
    chartC41.y_axis.majorGridlines = None
    chartC41.width = 15
    chartC41.height = 10
    chartC41.legend.position = 't'
    chartC41.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    sheetC.add_chart(chartC41, 'H71')

    chartC42 = PieChart()
    dataC42 = Reference(sheetC, min_col = 3, min_row = 91, max_col = 6)
    labelC42 = Reference(sheetC, min_col = 3, min_row = 70, max_col = 6)
    chartC42.add_data(dataC42, from_rows = 91, titles_from_data = False)
    chartC42.set_categories(labelC42)
    chartC42.width = 5
    chartC42.height = 5
    chartC42.legend = None
    chartC42.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chartC42.dLbls = DataLabelList()
    chartC42.dLbls.showPercent = True
    sheetC.add_chart(chartC42, 'M70')

    # if Data.rb == 1:
    #     wb.save('{}.xlsx'.format(savepath))
    #     tkinter.messagebox.showinfo('messagebox', '그래프 생성 완료(기술분류X)')

def D():
    sheetD = wb.create_sheet('기술분류별 출원동향', 3)
    D그래프data = Data.기술분류()
    for r in dataframe_to_rows(D그래프data, index=False, header=True):
        sheetD.append(r)

    sheetD.insert_cols(2)
    for row, cellobj in enumerate(list(sheetD.columns)[1]):
        n = '=right(A%d,2)' % (row+1)
        cellobj.value = n

    sheetD['B22'] = '합계'
    sheetD['C22'] = '=SUM(C2:C21)'
    sheetD['D22'] = '=SUM(D2:D21)'
    sheetD['E22'] = '=SUM(E2:E21)'
    sheetD['F22'] = '=SUM(F2:F21)'

    if Data.rb == 2:
        chartD11 = LineChart()
        dataD11 = Reference(sheetD, min_col = 3, min_row = 1, max_col = 4, max_row = 21)
        catsD11 = Reference(sheetD, min_col = 2, min_row = 2, max_row = 21)
        chartD11.add_data(dataD11, titles_from_data = True)
        chartD11.set_categories(catsD11)
        chartD11.y_axis.majorGridlines = None
        chartD11.width = 15
        chartD11.height = 10
        chartD11.legend.position = 't'
        chartD11.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
        sheetD.add_chart(chartD11, 'H2')
        
        chartD12 = PieChart()
        dataD12 = Reference(sheetD, min_col = 3, min_row = 22, max_col = 4)
        labelD12 = Reference(sheetD, min_col = 3, min_row = 1, max_col = 4)
        chartD12.add_data(dataD12, from_rows = 22, titles_from_data = False)
        chartD12.set_categories(labelD12)
        chartD12.width = 5
        chartD12.height = 5
        chartD12.legend = None
        chartD12.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
        chartD12.dLbls = DataLabelList()
        chartD12.dLbls.showPercent = True
        sheetD.add_chart(chartD12, 'M1')

        # wb.save('{}.xlsx'.format(savepath))
        # tkinter.messagebox.showinfo('messagebox', '그래프 생성 완료(기술분류 2개)')

    elif Data.rb == 3:
        chartD11 = LineChart()
        dataD11 = Reference(sheetD, min_col = 3, min_row = 1, max_col = 5, max_row = 21)
        catsD11 = Reference(sheetD, min_col = 2, min_row = 2, max_row = 21)
        chartD11.add_data(dataD11, titles_from_data = True)
        chartD11.set_categories(catsD11)
        chartD11.y_axis.majorGridlines = None
        chartD11.width = 15
        chartD11.height = 10
        chartD11.legend.position = 't'
        chartD11.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
        sheetD.add_chart(chartD11, 'H2')
        
        chartD12 = PieChart()
        dataD12 = Reference(sheetD, min_col = 3, min_row = 22, max_col = 5)
        labelD12 = Reference(sheetD, min_col = 3, min_row = 1, max_col = 5)
        chartD12.add_data(dataD12, from_rows = 22, titles_from_data = False)
        chartD12.set_categories(labelD12)
        chartD12.width = 5
        chartD12.height = 5
        chartD12.legend = None
        chartD12.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
        chartD12.dLbls = DataLabelList()
        chartD12.dLbls.showPercent = True
        sheetD.add_chart(chartD12, 'M1')

        # wb.save('{}.xlsx'.format(savepath))
        # tkinter.messagebox.showinfo('messagebox', '그래프 생성 완료(기술분류 3개)')

    elif Data.rb == 4:
        chartD11 = LineChart()
        dataD11 = Reference(sheetD, min_col = 3, min_row = 1, max_col = 6, max_row = 21)
        catsD11 = Reference(sheetD, min_col = 2, min_row = 2, max_row = 21)
        chartD11.add_data(dataD11, titles_from_data = True)
        chartD11.set_categories(catsD11)
        chartD11.y_axis.majorGridlines = None
        chartD11.width = 15
        chartD11.height = 10
        chartD11.legend.position = 't'
        chartD11.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
        sheetD.add_chart(chartD11, 'H2')
        
        chartD12 = PieChart()
        dataD12 = Reference(sheetD, min_col = 3, min_row = 22, max_col = 6)
        labelD12 = Reference(sheetD, min_col = 3, min_row = 1, max_col = 6)
        chartD12.add_data(dataD12, from_rows = 22, titles_from_data = False)
        chartD12.set_categories(labelD12)
        chartD12.width = 5
        chartD12.height = 5
        chartD12.legend = None
        chartD12.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
        chartD12.dLbls = DataLabelList()
        chartD12.dLbls.showPercent = True
        sheetD.add_chart(chartD12, 'M1')

        # wb.save('{}.xlsx'.format(savepath))
        # tkinter.messagebox.showinfo('messagebox', '그래프 생성 완료(기술분류 4개)')

def E():
    sheetE = wb.create_sheet('주요국 내외국인 출원점유율', 4)
    E그래프data1 = Data.내외국인점유율()
    for r in dataframe_to_rows(E그래프data1, index=False, header=True):
        sheetE.append(r)
    sheetE['A4'] = '합계'
    sheetE['B4'] = '=SUM(B2:B3)'
    sheetE['C4'] = '=SUM(C2:C3)'
    sheetE['D4'] = '=SUM(D2:D3)'
    sheetE['E4'] = '=SUM(E2:E3)'
    sheetE['A5'] = ' '

    E그래프data2 = Data.외국인점유율()
    for r in dataframe_to_rows(E그래프data2, index=False, header=True):
        sheetE.append(r)

    wb.save('{}.xlsx'.format(savepath))
    tkinter.messagebox.showinfo('messagebox', '그래프 생성 완료')